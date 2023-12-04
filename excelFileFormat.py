from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, GradientFill
import graphCreator


# Define color for init and end for gradient coloring of headers cells.
# They can be the same if only one color is wanted for this column / row with headers.
colorInitHeader = '7CAAF0'
colorEndHeader = '7CAAF0'


# Open Workbook and modify it.
def excel_rework(filename):
    wb = load_workbook(filename)
    print('Contents:')
    # Iterate through each worksheet.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(ws)
        # Set the columns' width according to text of each column
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            length = (length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = length
        # Gradient color for one of the cells.
        ws['B5'].fill = fill = GradientFill(stop=("FFFFFF", "000000"))

    wb.save(filename)


# Set format in the first row in which the headers are placed.
def format_header_horizontal(filename):
    wb = load_workbook(filename)
    # Iterate through each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply the desired format 1st column containing headers.
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader, fill_type='solid')
    wb.save(filename)


# Set format in the first column in which the headers are placed.
def format_header_vertical(filename):
    wb = load_workbook(filename)
    # Iterate through each worksheet.
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Apply the desired format 1st row containing headers.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader, fill_type='solid')
    wb.save(filename)


# Function to determine if the worksheet headers are horizontal or vertical.
# Assumption is to have more data than headers.
def headers_direction(ws):
    if (ws.max_row > ws.max_column) or ("Vertical" in ws.title):
        # Apply the desired format 1st column containing headers.
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            for cell in row:
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader, fill_type='solid')
    elif (ws.max_row < ws.max_column) or ("Horizontal" in ws.title):
        # Apply the desired format 1st row containing headers.
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = Font(name='Arial', size=12, bold=True)
                cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader, fill_type='solid')
    else:
        # Inconclusive result. All cases should be already be covered in the previous cases, but in the case we have a
        # squared matrix with headers in both vertical and horizontal axis, we can treat it here.
        if "Matrix" in ws.title:
            # Apply the desired format 1st column containing headers.
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(name='Arial', size=12, bold=True)
                    cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader, fill_type='solid')
            # Apply the desired format 1st row containing headers.
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = Font(name='Arial', size=12, bold=True)
                    cell.fill = PatternFill(start_color=colorInitHeader, end_color=colorEndHeader,
                                            fill_type='solid')
        pass


# Function to format header's cells in each according to format of the worksheet.
def format_header_specific(filename):
    wb = load_workbook(filename)
    # Iterate through each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers_direction(ws)
        wb.save(filename)


# Generic function for charts.
# Take as input parameter the type of chart to insert in the Excel file (to come) and the filename.
def chart(filename, type):
    wb = load_workbook(filename)
    wb.save()

