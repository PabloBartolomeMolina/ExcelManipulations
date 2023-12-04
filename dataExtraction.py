import csv
import openpyxl
import shutil
import graphCreator

csv_file = 'D:/Python_Projects/ExcelManipulations/airtravel.csv'
csv_file2 = 'D:/Python_Projects/ExcelManipulations/airtravel_horizontal.csv'


# Function to read a CSV and create an Excel ".xlsx" file with the contents.
def csv_to_excel(filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Rawdata'
    # CSV file can be downloaded from "https://people.sc.fsu.edu/~jburkardt/data/csv/csv.html".
    with open(csv_file) as f:
        reader = csv.reader(f, delimiter=',')   # Careful with the delimiter of your CSV.
        for row in reader:
            ws.append(row)
        wb.save(filename)
        # Include graph.
        chart(filename, ws, 0)  # Type = 0, for the moment only 1 type to test.
    ws = wb.create_sheet("Horizontal_Data")
    with open(csv_file2) as f:
        reader = csv.reader(f, delimiter=',')   # Careful with the delimiter of your CSV.
        for row in reader:
            ws.append(row)
    wb.save(filename)
    print(f'File created, {filename}')


# Function to read an Excel file and create a second one with the same contents and new name.
def rename_excel(input_file, output_file):
    try:
        # Copy the file without accessing its content.
        shutil.copyfile(input_file, output_file)
        print(f"File '{input_file}' successfully copied as '{output_file}'.")

    except Exception as e:
        print(f"An error occurred: {e}")


# Generic function for charts.
# Take as input parameter the type of chart to insert in the Excel file (to come) and the filename.
def chart(filename, ws, type):
    wb = openpyxl.load_workbook(filename)
    graphCreator.line_chart_create(wb, ws, filename, (2, 7, 2, 13), "Rawdata!B1:D1")
    wb.save(filename)
