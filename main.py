import openpyxl
import csv

from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, GradientFill, Alignment
import time

excel_file = 'D:/Python_Projects/ExcelManipulations/airtravel.xlsx'
csv_file = 'D:/Python_Projects/ExcelManipulations/airtravel.csv'


# Function to read a CSV and create an Excel ".xlsx" file with the contents.
def csv_to_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Raw data'
    # CSV file can be downloaded from "https://people.sc.fsu.edu/~jburkardt/data/csv/csv.html"
    with open(csv_file) as f:
        reader = csv.reader(f, delimiter=',')   # Careful with the delimiter of your CSV.
        for row in reader:
            ws.append(row)
    wb.save(excel_file)
    print(f'File created, {excel_file}')


# Open Workbook and modify it.
def excel_rework():
    wb = load_workbook(excel_file)
    ws = wb.active
    print('Contents:')
    print(ws)

    # Format for 1st row : Font size = 12 and Bold letters
    # Set a background color to this first row
    # Set centered alignment to this first row
    for column_cell in ws.iter_cols(1, 1):  # iterate column cells
        for row_cell in ws.iter_rows(1, ws.max_row + 1):
            for cell in column_cell:
                cell.font = Font(b=True, size=12)
                cell.fill = PatternFill(start_color="7CAAF0", end_color="7CAAF0", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set the columns' width according to text of each column
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        length = (length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = length

    # Gradient color for one of the cells.
    ws['A5'].fill = fill = GradientFill(stop=("FFFFFF", "000000"))

    wb.save(excel_file)

# Main entry point.
if __name__ == '__main__':
    csv_to_excel()
    excel_rework()
